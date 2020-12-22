from typing import Type

from openpyxl.cell import Cell
from openpyxl.styles import (
    Fill,
    PatternFill,
    Border,
    Side,
    Alignment,
    Protection,
    Font,
)


font = Font(
    name="Calibri",
    size=11,
    bold=False,
    italic=False,
    vertAlign=None,
    underline="none",
    strike=False,
    color="00FFFFFF",
)
fill = PatternFill(
    fill_type="solid",
    start_color="00339966",
)
required_fill = PatternFill(
    fill_type="solid",
    start_color="00800000",
)
border = Border(
    left=Side(border_style=None, color="FF000000"),
    right=Side(border_style=None, color="FF000000"),
    top=Side(border_style=None, color="FF000000"),
    bottom=Side(border_style=None, color="FF000000"),
    diagonal=Side(border_style=None, color="FF000000"),
    diagonal_direction=0,
    outline=Side(border_style=None, color="FF000000"),
    vertical=Side(border_style=None, color="FF000000"),
    horizontal=Side(border_style=None, color="FF000000"),
)
alignment = Alignment(
    horizontal="general",
    vertical="bottom",
    text_rotation=0,
    wrap_text=False,
    shrink_to_fit=False,
    indent=0,
)
number_format = "General"
protection = Protection(locked=True, hidden=False)


class Style:
    def __init__(
        self,
        font: Type[Font],
        fill: Type[Fill],
        border: Type[Border],
        alignment: Type[Alignment],
        number_format: str,
        protection: Type[Protection],
    ):
        self.font = font
        self.fill = fill
        self.border = border
        self.alignment = alignment
        self.number_format = number_format
        self.protection = protection


def apply_style(cell: Type[Cell], style: Type[Style]):
    cell.font = style.font
    cell.fill = style.fill
    cell.border = style.border
    cell.alignment = style.alignment
    cell.number_format = style.number_format
    cell.protection = style.protection


header_style = Style(
    font=font,
    fill=fill,
    border=border,
    alignment=alignment,
    number_format=number_format,
    protection=protection,
)

required_header_style = Style(
    font=font,
    fill=required_fill,
    border=border,
    alignment=alignment,
    number_format=number_format,
    protection=protection,
)